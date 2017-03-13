# -*- coding:utf8 -*-
import urllib,urllib2
import time,sys
import re
from openpyxl import load_workbook


def transCorpus(str_cps):
    url_get_base = "http://api.ltp-cloud.com/analysis/"
    # print str_cps
    args = {
        'api_key' : 'C8G6c1l5ZzLI3SWrPm4TQPPVLHmpCATEffRQWxxt',
        'text' : str_cps.encode('utf8'),
        'pattern' : 'pos',
        'format' : 'plain'
    }
    try:
        result = urllib.urlopen(url_get_base, urllib.urlencode(args)) # POST method
        #  content = result.read().strip()
        #  print result
        return result.read()
    except urllib2.HTTPError, e:
        print >> sys.stderr, e.reason


#open xls to be checked
wb_corpus = load_workbook('corpus.xlsx')
ws_corpus = wb_corpus.get_sheet_by_name("ltp")

r = 1
while (r <= ws_corpus.max_row):
    str_corpus = ws_corpus.cell(row=r, column=1).value
    print "原始语料：", str_corpus
    result_pos = transCorpus(str_corpus)
    #语言云限制频率200次/s
    time.sleep(0.2)
    if result_pos.strip() == '':
        r += 0
    else:
        # print "词性解析结果：", result_pos
        ws_corpus.cell(row=r, column=2).value = result_pos
        # pos_list = re.search('(.+(\_(a-n|nd|nh|ni|nl|ns|nt|nz|o-r|u|v|wp|ws|x)\s){1})+', str_pos)
        re_pos = re.compile(r'\_[a-z]+\s?')
        pos_list = re_pos.findall(result_pos)
        print "词性序列：", pos_list
        str_pos = "".join(pos_list)
        ws_corpus.cell(row=r, column=3).value = str_pos
        r += 1

wb_corpus.save('corpus.xlsx')

